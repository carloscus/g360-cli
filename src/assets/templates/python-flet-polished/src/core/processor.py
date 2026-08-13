from __future__ import annotations

"""
G360 Processor - Motor de logica de negocio.

Este modulo debe ser heredado y extendido por cada app G360.
Contiene los patrones base para:
- Calcular KPIs agregados
- Generar metricas por linea/categoria
- Sugiere transferencias entre almacenes
- Exporta reportes a Excel con formato profesional

Heredar y sobrescribir los metodos segun el dominio de cada app.
"""

from datetime import datetime
from pathlib import Path


APP_AUTHOR = "g360-app-polished"
APP_NAME = "G360"


def _make_report_name(title: str) -> str:
    """Genera nombre de archivo con timestamp: G360_{slug}_{YYYYMMDD}_{HHMMSS}.xlsx"""
    now = datetime.now()
    ts = now.strftime("%Y%m%d_%H%M%S")
    import re
    slug = re.sub(r"[^a-zA-Z0-9]+", "_", title).strip("_").upper()
    if len(slug) > 40:
        slug = slug[:40]
    if not slug:
        slug = "REPORTE"
    return f"{APP_NAME}_{slug}_{ts}"


class BaseProcessor:
    """
    Procesador base que debe heredarse.

    Ejemplo de uso:
        class MiProcessor(BaseProcessor):
            def calcular_kpis(self, raw_data):
                # Implementar logica especifica
                return kpis_dict
    """

    def __init__(self):
        self._last_kpis = None
        self._last_hash = None

    def calcular_kpis(self, raw_data: dict) -> dict:
        """
        Calcula KPIs a partir de datos crudos.
        Debe ser sobrescrito por subclases.

        Args:
            raw_data: Datos crudos del dominio (dict por almacén/entidad)

        Returns:
            dict con KPIs calculados
        """
        raise NotImplementedError("Subclase debe implementar calcular_kpis()")

    def obtener_metricas(self, kpis: dict) -> tuple[list[dict], list[dict]]:
        """
        Genera metricas agrupadas (por linea, categoria, etc).
        Debe ser sobrescrito por subclases.

        Returns:
            (metricas_principales, metricas_secundarias)
        """
        raise NotImplementedError("Subclase debe implementar obtener_metricas()")

    def sugerir_acciones(self, raw_data: dict, kpis: dict) -> list[dict]:
        """
        Sugiere acciones (transferencias, alertas, etc).
        Debe ser sobrescrito por subclases.

        Returns:
            lista de dicts con sugerencias
        """
        return []

    def export_to_excel(self, data: list, file_path: str, title: str = "Reporte"):
        """
        Exporta datos a Excel con formato profesional.
        Metodo base reutilizable.

        Args:
            data: Lista de filas [(col1, col2, ...), ...]
            file_path: Ruta del archivo de salida
            title: Titulo del reporte
        """
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Font, Alignment

        wb = Workbook()
        wb.properties.creator = APP_AUTHOR
        wb.properties.description = f"Reporte generado por {APP_NAME} — {datetime.now().strftime('%Y-%m-%d %H:%M')}"

        ws = wb.active
        ws.title = title[:31]

        # Headers
        if data:
            headers = [f"Col {i+1}" for i in range(len(data[0]))]
            ws.append(headers)
            for cell in ws[1]:
                cell.fill = PatternFill(start_color="10B981", end_color="10B981", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
                cell.alignment = Alignment(horizontal="center")

            for row in data:
                ws.append([str(v) if v is not None else "" for v in row])

        ws.column_dimensions['A'].width = 35
        for col in "BCD":
            ws.column_dimensions[col].width = 16

        wb.save(file_path)

    def hash_data(self, data: dict) -> str:
        """Hash SHA-256 para deteccion de cambios."""
        import hashlib
        import json
        serialized = json.dumps(data, sort_keys=True, ensure_ascii=False).encode()
        return hashlib.sha256(serialized).hexdigest()[:16]
